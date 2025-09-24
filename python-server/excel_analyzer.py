#!/usr/bin/env python3
"""
Excel Formula Analyzer for Ternus Pricing Engine
Extracts formulas, cell references, and structure from Excel template
"""

import openpyxl
import pandas as pd
import json
from typing import Dict, List, Any
import re
import os

class ExcelFormulaAnalyzer:
    def __init__(self, excel_path: str):
        self.excel_path = excel_path
        self.workbook = None
        self.formulas = {}
        self.cell_values = {}
        self.worksheets = {}
        
    def load_excel(self):
        """Load the Excel file"""
        try:
            self.workbook = openpyxl.load_workbook(self.excel_path, data_only=False)
            print(f"✅ Loaded Excel file: {self.excel_path}")
            print(f"📊 Worksheets: {self.workbook.sheetnames}")
            return True
        except Exception as e:
            print(f"❌ Error loading Excel file: {e}")
            return False
    
    def extract_formulas(self):
        """Extract all formulas from all worksheets"""
        if not self.workbook:
            return
        
        for sheet_name in self.workbook.sheetnames:
            sheet = self.workbook[sheet_name]
            sheet_formulas = {}
            sheet_values = {}
            
            print(f"\n🔍 Analyzing sheet: {sheet_name}")
            
            # Get sheet dimensions
            max_row = sheet.max_row
            max_col = sheet.max_column
            print(f"📏 Sheet size: {max_row} rows x {max_col} columns")
            
            formula_count = 0
            value_count = 0
            
            for row in range(1, max_row + 1):
                for col in range(1, max_col + 1):
                    cell = sheet.cell(row=row, column=col)
                    cell_ref = f"{openpyxl.utils.get_column_letter(col)}{row}"
                    
                    # Extract formulas
                    if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                        sheet_formulas[cell_ref] = {
                            'formula': cell.value,
                            'row': row,
                            'col': col,
                            'column_letter': openpyxl.utils.get_column_letter(col)
                        }
                        formula_count += 1
                    
                    # Extract values (including calculated values)
                    elif cell.value is not None:
                        sheet_values[cell_ref] = {
                            'value': cell.value,
                            'data_type': type(cell.value).__name__,
                            'row': row,
                            'col': col,
                            'column_letter': openpyxl.utils.get_column_letter(col)
                        }
                        value_count += 1
            
            self.formulas[sheet_name] = sheet_formulas
            self.cell_values[sheet_name] = sheet_values
            
            print(f"📋 Found {formula_count} formulas and {value_count} values")
    
    def analyze_formula_patterns(self):
        """Analyze formula patterns to understand the pricing logic"""
        print("\n🔬 FORMULA ANALYSIS")
        print("=" * 50)
        
        all_formulas = []
        
        for sheet_name, sheet_formulas in self.formulas.items():
            print(f"\n📊 Sheet: {sheet_name}")
            print("-" * 30)
            
            for cell_ref, formula_data in sheet_formulas.items():
                formula = formula_data['formula']
                all_formulas.append(formula)
                
                print(f"  📍 {cell_ref}: {formula}")
                
                # Analyze formula components
                self._analyze_formula_components(formula, cell_ref)
        
        # Find common patterns
        self._find_common_patterns(all_formulas)
    
    def _analyze_formula_components(self, formula: str, cell_ref: str):
        """Analyze individual formula components"""
        # Find cell references
        cell_refs = re.findall(r'[A-Z]+\d+', formula)
        if cell_refs:
            print(f"    🔗 References: {', '.join(set(cell_refs))}")
        
        # Find functions
        functions = re.findall(r'([A-Z]+)\(', formula)
        if functions:
            print(f"    ⚡ Functions: {', '.join(set(functions))}")
        
        # Find constants
        constants = re.findall(r'\d+\.?\d*', formula)
        if constants:
            print(f"    🔢 Constants: {', '.join(set(constants))}")
    
    def _find_common_patterns(self, formulas: List[str]):
        """Find common patterns across all formulas"""
        print(f"\n📈 COMMON PATTERNS ANALYSIS")
        print("-" * 30)
        
        # Count functions
        function_counts = {}
        for formula in formulas:
            functions = re.findall(r'([A-Z]+)\(', formula)
            for func in functions:
                function_counts[func] = function_counts.get(func, 0) + 1
        
        print("🔧 Most used functions:")
        for func, count in sorted(function_counts.items(), key=lambda x: x[1], reverse=True):
            print(f"  {func}: {count} times")
    
    def export_analysis(self, output_file: str = "pricing_analysis.json"):
        """Export analysis results to JSON"""
        analysis_data = {
            'worksheets': list(self.workbook.sheetnames) if self.workbook else [],
            'formulas': self.formulas,
            'cell_values': self.cell_values,
            'summary': {
                'total_formulas': sum(len(sheet) for sheet in self.formulas.values()),
                'total_values': sum(len(sheet) for sheet in self.cell_values.values()),
                'sheets_count': len(self.formulas)
            }
        }
        
        with open(output_file, 'w') as f:
            json.dump(analysis_data, f, indent=2, default=str)
        
        print(f"\n💾 Analysis exported to: {output_file}")
        return analysis_data
    
    def generate_pricing_config(self):
        """Generate pricing engine configuration from Excel analysis"""
        config = {
            'pricing_rules': {},
            'input_variables': [],
            'output_calculations': [],
            'lookup_tables': {},
            'formulas': {}
        }
        
        for sheet_name, sheet_formulas in self.formulas.items():
            config['formulas'][sheet_name] = {}
            
            for cell_ref, formula_data in sheet_formulas.items():
                formula = formula_data['formula']
                
                # Convert Excel formula to Python-compatible format
                python_formula = self._convert_excel_to_python(formula)
                
                config['formulas'][sheet_name][cell_ref] = {
                    'excel_formula': formula,
                    'python_formula': python_formula,
                    'dependencies': re.findall(r'[A-Z]+\d+', formula)
                }
        
        return config
    
    def _convert_excel_to_python(self, formula: str) -> str:
        """Convert Excel formula to Python-compatible format"""
        # Remove leading =
        python_formula = formula[1:] if formula.startswith('=') else formula
        
        # Convert Excel functions to Python equivalents
        conversions = {
            'SUM': 'sum',
            'MAX': 'max',
            'MIN': 'min',
            'IF': 'np.where',
            'VLOOKUP': 'vlookup',  # Custom function
            'PMT': 'pmt',          # Custom function
            'POWER': 'pow',
            'SQRT': 'math.sqrt',
            'ABS': 'abs',
            'ROUND': 'round'
        }
        
        for excel_func, python_func in conversions.items():
            python_formula = re.sub(f'{excel_func}\\(', f'{python_func}(', python_formula)
        
        return python_formula


def main():
    """Main function to analyze the Excel file"""
    excel_file = "../TERM SHEET TEMPLATE V 6.9.2025.xltx"
    
    if not os.path.exists(excel_file):
        print(f"❌ Excel file not found: {excel_file}")
        return
    
    print("🚀 STARTING EXCEL ANALYSIS FOR PRICING ENGINE")
    print("=" * 60)
    
    analyzer = ExcelFormulaAnalyzer(excel_file)
    
    # Load and analyze
    if analyzer.load_excel():
        analyzer.extract_formulas()
        analyzer.analyze_formula_patterns()
        
        # Export results
        analysis_data = analyzer.export_analysis()
        pricing_config = analyzer.generate_pricing_config()
        
        # Save pricing config
        with open("pricing_config.json", 'w') as f:
            json.dump(pricing_config, f, indent=2, default=str)
        
        print(f"\n🎯 PRICING ENGINE READY!")
        print(f"📊 Found {analysis_data['summary']['total_formulas']} formulas")
        print(f"📋 Found {analysis_data['summary']['total_values']} values")
        print(f"📝 Configuration saved to: pricing_config.json")
        
        return analysis_data, pricing_config
    
    return None, None


if __name__ == "__main__":
    main() 